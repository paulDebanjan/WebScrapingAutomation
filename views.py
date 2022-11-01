from django.http import HttpResponse
from msilib.schema import File
from django.shortcuts import render
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook,load_workbook
from django.conf import settings
from django.core.files import File
import os
from django.utils.decorators import method_decorator
from ..userAuthentication.decorators import employee_required
from django.contrib.auth.decorators import login_required

@login_required
@employee_required
def searchIndex(request):
    instagram = ''
    linkedin = ''
    facebook = ''
    twitter = ''

    if request.method == 'POST':
        url = request.POST.get('url_text')
        html_text = requests.get(url).text
        soup = BeautifulSoup(html_text,'lxml')
        company_info = soup.find_all('h3', class_="company_info")

        workbook = Workbook()
        worksheet = workbook.active

        for company in company_info:
            link = company.a['href']
            worksheet.append([link])

        workbook['Sheet'].title = "Content_Link"
        workbook.save("report.xlsx")

        
        wb = load_workbook('report.xlsx')
        sheet = wb['Content_Link']
        row = sheet.max_row
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Company Name","Title",'Address',"Image Link",'Official Website','Teliphone',"description",'Avg Hourly Rate','Employees','Min Project Size','Founded','Facebook','Linkedin','Twitter','Instagram'])

        for i in range(1, row + 1):
            html_text = requests.get("https://clutch.co/" + sheet.cell(row = i, column = 1).value).text
            soup = BeautifulSoup(html_text, 'lxml')
            company_header = soup.find('div', class_='header-company company_logotype')                                 # Select Header Dev
            company_image = company_header.a.img['src']                                                                 # Collect Company Logo
            Company_name = company_header.find('h1', class_='header-company--title').text.strip()                       # Collect Company Name
            company_official_link = company_header.h1.a['href']                                                         # Collect Company Official Address
            company_description_dev = soup.find('div', class_='summary-description')                                    # Select Company Dev
            title = company_description_dev.find('h2', class_='h2_title').text
            company_description = company_description_dev.find('div', class_='field-name-profile-summary').p.text       # Collect Description About Company
            short_infomation_div = company_description_dev.find_all('div', class_='list-item custom_popover')
            for i in short_infomation_div:
                if i['data-content'] == "<i>Avg. hourly rate</i>":
                    avgHourlyRate = i.text.strip()
                elif i['data-content'] == "<i>Employees</i>":
                    employees = i.text.strip()
                elif i['data-content'] == "<i>Founded</i>":
                    founded = i.text.strip()
            minProjectSizeDiv = company_description_dev.find('div', class_='list-item custom_popover custom_popover__left')
            if minProjectSizeDiv['data-content'] == "<i>Min. project size</i>":
                minProjectSize = minProjectSizeDiv.text.strip()
            address_div = soup.find('ul', class_='location-address headquarters')
            address1 = address_div.li.div.div.div.find('div', class_="street-address").text.strip()
            address2 = address_div.li.div.div.div.find('span', class_="locality").text.strip()
            address3 = address_div.li.div.div.div.find('span', class_="region").text.strip()
            address4 = address_div.li.div.div.div.find('span', class_="postal-code").text.strip()
            address5 = address_div.li.div.div.div.find('div', class_="country-name").text.strip()
            address = address1 + ", " +address2 + ', ' + address3 + ', ' + address4 +', ' + address5

            teliphone = address_div.li.div.div.div.a['href'].replace("%"," ")
            socila_medial_group = soup.find("li", class_="profile-social-wrap")                             # Selecting Social upper tag
            if socila_medial_group is not None: 
                socila_media_list = socila_medial_group.find_all("a",class_='profile-social-link')
                for k in socila_media_list:
                    if k['data-type'] == 'facebook':                                                    # Matching With Facebook Keyword
                        facebook = k['href']
                    elif k['data-type'] == 'linkedin':                                                  # Matching With Facebook Linkedin
                        linkedin = k['href']
                    elif k['data-type'] == 'twitter':                                                   # Matching With Facebook twitter
                        twitter = k['href']
                    elif k['data-type'] == 'instagram':                                                 # Matching With Facebook Instagram
                        instagram = k['href']
            worksheet.append([Company_name,title,address,company_image,company_official_link,teliphone,company_description,avgHourlyRate,employees,minProjectSize,founded,facebook,linkedin,twitter,instagram]) 
            instagram = ''
            linkedin = ''
            facebook = ''
            twitter = ''
        workbook.save('finalReport.xlsx')
        with open(f'{settings.BASE_DIR}/finalReport.xlsx', 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(f'{settings.BASE_DIR}/finalReport.xlsx')
        return response
        # return render(request, 'dataCapture/result.html',{'user_data': user_data})
        
    return render(request,"dataCapture/search.html")